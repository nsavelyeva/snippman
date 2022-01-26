import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from .messages import msg
from . import helpers


def save_to_local_file(rows, app_conf):
    w_book = Workbook()
    w_sheet = w_book.active
    w_sheet.title = "Snippets"
    title_row = ["Language", "Description", "Snippet", "Comment",
                 "Created", "Modified", "Accessed", "Visits"]
    w_sheet.append(title_row)
    for row in rows:
        # Skip first column returned by join-query in db_queries.get_all_snippets():
        w_sheet.append(row[1:])
    file_name = "snippets_%s.xlsx" % datetime.now().strftime("%Y-%m-%d_%H-%M-%S-%f")[:-3]
    folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), app_conf["DOWNLOAD_FOLDER"])
    link = "%s/%s" % (folder, file_name)
    errors = []
    try:
        w_book.save(filename=link)
        errors.extend(helpers.delete_old_local_excels(app_conf).errors)  # Remove old backups
    except IOError as err:
        errors.append(msg(1041, "'%s' due to %s" % (file_name, err)))    # Failed to save file
        return helpers.Data(None, errors)
    return helpers.Data(file_name, [])


def read_from_local_file(file_name):
    try:
        w_book = load_workbook(file_name)
    except InvalidFileException as err:
        return helpers.Data(None, [msg(1051, "'%s' due to %s" % (file_name, err))])  # import failed
    w_sheet = w_book.get_sheet_by_name("Snippets")
    result = enumerate(w_sheet.iter_rows())
    return helpers.Data(result, [])
